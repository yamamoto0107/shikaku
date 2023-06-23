"""
要件定義
・前回作成した答え合わせソフトで、画面設計。
・openpyxlとファイル選択機能。
・フォルダでの一括添削機能。
・入力欄１：正解エクセルファイルの選択。
・入力欄２：回答エクセルファイルの選択。
・入力欄３：正誤を〇×形式か〇答え形式選択。
・入力欄４：添削ファイルの保存場所設定。(初期値は回答エクセルのアドレス)
・入力欄５：学生の回答問題数と回答開始行と列番号。
・出力欄１：処理の結果(成功か失敗か。)
・出力欄２：点数。
(初期値は、20,1,2)

・機能１：エクセルファイルの正誤入力と保存。
・機能２：正答率の表示。
・機能３：ファイル選択画面。
"""
import openpyxl as xl
import subprocess
import sys
import os
try:
    from PyQt5 import QtWidgets,QtGui
    from PyQt5.QtWidgets import QApplication, QWidget,QFileDialog,QListWidget,QSizePolicy,QVBoxLayout,QHBoxLayout
except:
    subprocess.call("python -m pip install pyqt5==5.15.9")
    print("もう一度実行してください。Enterで終了")
    x=input()
    sys.exit(0)

class Main(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
    def initUI(self):
        self.setWindowTitle("")
        self.resize(640,480)
        self.layout4 = QtWidgets.QFormLayout()
        self.label()
        self.line()
        self.file()  
        self.setLayout(self.layout4)
        self.layout4.addRow("答え合わせする問題数を教えて下さい。", self.line1)
        self.layout4.addRow("正解ファイルを選択してください。", self.button)
        self.layout4.addRow(self.label1)
        self.layout4.addRow("回答ファイルを選択してください。", self.button1)
        self.layout4.addRow(self.label2)
        self.layout4.addRow("回答開始行を教えて下さい。", self.line2)
        self.layout4.addRow("回答の書かれた列を教えて下さい。", self.line3)
        self.layout4.addRow(self.label0)
    def label(self):
        self.label0 = QtWidgets.QLabel(self)
        self.label1 = QtWidgets.QLabel(self)
        self.label2 = QtWidgets.QLabel(self)
        self.label0.setFont(QtGui.QFont("Arial", 24))
        self.label1.setText("\t\t\t\t上のボタンから正解のエクセルファイルを選択してください")
        self.label2.setText("\t\t\t\t上のボタンから学生のエクセルファイルを選択してください")
        self.label0.setText("各データを入力してください。")
    def line(self):
        self.line1 = QtWidgets.QLineEdit(self)
        self.line1.setFixedWidth(150)
        self.line1.setText("13")
        self.line2 = QtWidgets.QLineEdit(self)
        self.line2.setFixedWidth(150)
        self.line2.setText("1")
        self.line3 = QtWidgets.QLineEdit(self)
        self.line3.setFixedWidth(150)
        self.line3.setText("2")
    
    def file(self):
        self.button = QtWidgets.QPushButton(self)
        self.button.setMinimumSize(320,100)
        self.button.setFixedWidth(320)
        self.button.setText("正解ファイルを選択する")
        self.button.clicked.connect(self.dialog_test)
        self.button1 = QtWidgets.QPushButton(self)
        self.button1.setMinimumSize(320,100)
        self.button1.setFixedWidth(320)
        self.button1.setText("回答ファイルを選択する")
        self.button1.clicked.connect(self.dialog_test1)

    def dialog_test(self):
        self.file0,_ = QFileDialog.getOpenFileName()
        self.label1.setText(self.file0+"が選択されました")
    def dialog_test1(self):
        self.file1,_ = QFileDialog.getOpenFileName()
        self.label2.setText(self.file1+"が選択されました")
        test=self.checker(self.file0,self.file1,self.file1,int(self.line1.text()),int(self.line2.text()),int(self.line3.text()))
        self.label0.setText("入力を確認しました。"+test+"ファイルに記入しましたので返却してください")

    def checker(self,ch,ans,save,row,start,col):
        try:
            wb1 = xl.load_workbook(ch)
            wb2 = xl.load_workbook(ans)
            ws1 = wb1.worksheets[0]
            ws2 = wb2.worksheets[0]
            countall=0
            count=0
            for i in range(row):#0+start行目から問題数+start行目まで
                countall+=1
                z=i+start
                cell1 = ws1.cell(z,col)
                cell2 = ws2.cell(z,col)
                if cell1.value == cell2.value:
                    print(cell2.value,"○")
                    ws2.cell(z,col+1,"○")
                    count+=1
                else:
                    ws2.cell(z,col+1,cell2.value)
            parsent=str(count//countall*100)+"点でした。"
            ws2.cell(row+1,col+1,parsent)
            wb2.save(save)

            return parsent
        except Exception as e:

            return e

ch = "C:/Users/yyama/Downloads/shikaku/答え合わせソフト/FE_0623ans.xlsx"
ans = "C:/Users/yyama/Downloads/shikaku/答え合わせソフト/FE_0623.xlsx"

if __name__ == "__main__":
    #正解、回答、出力先、問題数、開始行、回答列
    app=QApplication(sys.argv)
    windows=Main()
    windows.show()
    sys.exit(app.exec_())
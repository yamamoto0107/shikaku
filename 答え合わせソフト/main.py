"""
要件定義
・前回作成した答え合わせソフトで、画面設計。
・openpyxlとファイル選択機能。
・フォルダでの一括添削機能。
・入力欄１：正解エクセルファイルの選択。
・入力欄２：回答エクセルファイルの選択。
・入力欄３：正誤を〇×形式か〇答え形式選択。
・入力欄４：添削ファイルの保存場所設定。(初期値は回答エクセルのアドレス)
・入力欄５：学生の回答問題数と回答開始行と列番号。
・出力欄１：処理の結果(成功か失敗か。)
・出力欄２：点数。
(初期値は、20,1,2)

・機能１：エクセルファイルの正誤入力と保存。
・機能２：正答率の表示。
・機能３：ファイル選択画面。
"""
import openpyxl as xl
import subprocess
import sys
import os
try:
    from PyQt5 import QtWidgets,QtGui
    from PyQt5.QtWidgets import QApplication, QWidget,QFileDialog,QListWidget,QSizePolicy,QVBoxLayout,QHBoxLayout
except:
    subprocess.call("python -m pip install pyqt5==5.15.9")
    print("もう一度実行してください。Enterで終了")
    x=input()
    sys.exit(0)

class Main(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
    def initUI(self):
        self.setWindowTitle("")
        self.resize(640,480)
        self.layout4 = QtWidgets.QFormLayout()
        self.label()
        self.line()
        self.file()  
        self.setLayout(self.layout4)
        self.layout4.addRow("答え合わせする問題数を教えて下さい。", self.line1)
        self.layout4.addRow("正解ファイルを選択してください。", self.button)
        self.layout4.addRow(self.label1)
        self.layout4.addRow("回答ファイルを選択してください。", self.button1)
        self.layout4.addRow(self.label2)
        self.layout4.addRow("回答開始行を教えて下さい。", self.line2)
        self.layout4.addRow("回答の書かれた列を教えて下さい。", self.line3)
        self.layout4.addRow(self.label0)
    def label(self):
        self.label0 = QtWidgets.QLabel(self)
        self.label1 = QtWidgets.QLabel(self)
        self.label2 = QtWidgets.QLabel(self)
        self.label0.setFont(QtGui.QFont("Arial", 24))
        self.label1.setFont(QtGui.QFont("Arial", 14))
        self.label2.setFont(QtGui.QFont("Arial", 14))
        self.label1.setText("\t\t上のボタンから正解のエクセルファイルを選択してください")
        self.label2.setText("\t\t上のボタンから学生のエクセルファイルを選択してください")
        self.label0.setText("各データを入力してください。")
    def line(self):
        self.line1 = QtWidgets.QLineEdit(self)
        self.line1.setFixedWidth(150)
        self.line1.setText("20")
        self.line2 = QtWidgets.QLineEdit(self)
        self.line2.setFixedWidth(150)
        self.line2.setText("2")
        self.line3 = QtWidgets.QLineEdit(self)
        self.line3.setFixedWidth(150)
        self.line3.setText("2")
    
    def file(self):
        self.button = QtWidgets.QPushButton(self)
        self.button.setMinimumSize(320,100)
        self.button.setFixedWidth(320)
        self.button.setText("正解ファイルを選択する")
        self.button.clicked.connect(self.dialog_test)
        self.button1 = QtWidgets.QPushButton(self)
        self.button1.setMinimumSize(320,100)
        self.button1.setFixedWidth(320)
        self.button1.setText("回答ファイルを選択する")
        self.button1.clicked.connect(self.dialog_test1)

    def dialog_test(self):
        self.file0,_ = QFileDialog.getOpenFileName()
        self.label1.setText(self.file0+"が選択されました")
    def dialog_test1(self):
        self.file1,_ = QFileDialog.getOpenFileName()
        self.label2.setText(self.file1+"が選択されました")
        try:
            test=self.checker(self.file0,self.file1,self.file1,int(self.line1.text()),int(self.line2.text()),int(self.line3.text()))
            text="入力を確認しました。"+self.parsent+"ファイルに記入しましたので返却してください"
            self.label0.setText(text)
        except Exception as e:
            text="もう一度、正解ファイルと回答ファイルを上から選択しなおしてください"
            self.label0.setText(text)
    def checker(self,ch,ans,save,row,start,col):
        try:
            wb1 = xl.load_workbook(ch)
            wb2 = xl.load_workbook(ans)
            ws1 = wb1.worksheets[0]
            ws2 = wb2.worksheets[0]
            countall=0
            count=0
            for i in range(row):#0+start行目から問題数+start行目まで
                countall+=1
                z=i+start
                cell1 = ws1.cell(z,col)
                cell2 = ws2.cell(z,col)
                if cell1.value == cell2.value:
                    print(cell2.value,"○")
                    ws2.cell(z,col+1,"○")
                    count+=1
                else:
                    ws2.cell(z,col+1,cell1.value)
                    print(cell1.value,cell2.value)
            answer=count/countall*100
            print(count,countall,int(answer))
            answer=int(answer)
            self.parsent=str(answer)+"点でした。"
            ws2.cell(row+2,col+1,self.parsent)
            wb2.save(save)

            return self.parsent
        except Exception as e:
            
            return e

ch = "./答え合わせソフト/回答用紙.xlsx"
ans = "./答え合わせソフト/正解データ.xlsx"

if __name__ == "__main__":
    #正解、回答、出力先、問題数、開始行、回答列
    app=QApplication(sys.argv)
    windows=Main()
    windows.show()
    sys.exit(app.exec_())